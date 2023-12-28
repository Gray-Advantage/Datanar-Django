from io import BytesIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic.detail import DetailView
from django.views.generic.list import ListView
import openpyxl

from redirects.models import Redirect
from statistic.models import Click


def get_clicks_by_mode(short_link, mode):
    if mode in ["year", "month", "day"]:
        return eval(
            f"Click.objects.for_short_link_by_last_{mode}('{short_link}')",
            {"Click": Click},
        )
    return Click.objects.for_short_link_by_all_time(short_link)


class MyLinksView(LoginRequiredMixin, ListView):
    template_name = "statistic/my_links.html"
    context_object_name = "links"

    def get_queryset(self):
        redirects = Redirect.objects.filter(user=self.request.user).only(
            Redirect.short_link.field.name,
        )
        return [redirect.short_link for redirect in redirects]


class LinkDetailView(LoginRequiredMixin, DetailView):
    template_name = "statistic/link_detail.html"
    context_object_name = "redirect"

    def get_object(self, queryset=None):
        return get_object_or_404(Redirect, short_link=self.kwargs["link"])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["link"] = self.kwargs["link"]
        context["period"] = self.kwargs["period"]

        clicks = get_clicks_by_mode(
            context["redirect"].short_link,
            self.kwargs["period"],
        )

        context.update(
            {
                "browser": self._get_statistic(
                    clicks,
                    Click.browser.field.name,
                ),
                "os": self._get_statistic(clicks, Click.os.field.name),
                "country": self._get_statistic(
                    clicks,
                    Click.country.field.name,
                ),
                "city": self._get_statistic(clicks, Click.city.field.name),
                "clicks": clicks.count(),
            },
        )

        return context

    def _get_statistic(self, clicks, field):
        res = {}
        for click in clicks:
            res[click.__dict__[field]] = res.get(click.__dict__[field], 0) + 1
        return res


class DownloadStatistic(View):
    def get(self, request, link, period):
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = [
            "#",
            _("Created"),
            _("OS"),
            _("Browser"),
            _("Country"),
            _("City"),
        ]
        headers = list(map(str, headers))

        for col_num, header in enumerate(headers, 1):
            ws[f"{openpyxl.utils.get_column_letter(col_num)}1"] = header

        for number, click in enumerate(get_clicks_by_mode(link, period), 1):
            ws.append(
                [
                    number,
                    click.clicked_at.replace(tzinfo=None),
                    click.os,
                    click.browser,
                    click.country,
                    click.city,
                ],
            )

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells) + 2
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(column_cells[0].column)
            ].width = length

        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)

        return FileResponse(
            virtual_workbook,
            as_attachment=True,
            filename="statistic.xlsx",
        )


__all__ = [MyLinksView]
